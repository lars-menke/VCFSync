"""
E-Mail aufräumen - Bewertung, Excel-Rundlauf und Ausführung
============================================================
Durchsucht ausgewählte IMAP-Postfächer und -Ordner, bewertet jede Mail nach
festen Regeln und schlägt begründet vor, was in den Papierkorb kann. Der
Ablauf ist bewusst derselbe wie beim Kontakte-Sync:

    scan  ->  to-excel  ->  (in Excel oder im Browser prüfen)  ->  from-excel
          ->  clean (Testlauf)  ->  clean --execute

Grundsätze:
  * **Nichts wird endgültig gelöscht.** Mails wandern in den Papierkorb des
    Kontos und können dort zurückgeholt werden.
  * **Der Testlauf ist der Standard.** Ein echter Lauf braucht --execute.
  * **Nur Kopfzeilen werden gelesen**, keine Mail-Texte. Alles bleibt lokal,
    es wird nichts an einen Dienst gesendet.
  * **Das Tool lernt mit**: Nach jedem echten Lauf merkt es sich, welche
    Absender du gelöscht und welche du bewusst behalten hast, und bewertet
    beim nächsten Mal entsprechend.

Zugangsdaten stehen in mail_accounts.json (nicht im Git). Für iCloud und
Gmail wird ein app-spezifisches Passwort benötigt, nicht das normale.
"""

import argparse
import json
import sys
from datetime import datetime, timezone
from getpass import getpass
from pathlib import Path

import mail_imap as mi
from icloud_contacts import _cell_is_delete

ACCOUNTS_FILE = Path("mail_accounts.json")
DECISIONS_FILE = Path("mail_decisions.json")
RULES_FILE = Path("mail_rules.json")
SCAN_FILE = Path("web_data") / "mail_scan.json"

# Die vier von Hand pflegbaren Regellisten.
RULE_LISTS = ("never_delete", "always_delete", "subject_never", "subject_delete")

# Bekannte Anbieter - erspart das Nachschlagen der Serveradressen.
PRESETS = {
    "icloud": {"host": "imap.mail.me.com", "port": 993},
    "gmail": {"host": "imap.gmail.com", "port": 993},
    "gmx": {"host": "imap.gmx.net", "port": 993},
    "web.de": {"host": "imap.web.de", "port": 993},
    "outlook": {"host": "outlook.office365.com", "port": 993},
}

DEFAULT_SETTINGS = {
    "min_age_days": 30,          # jünger wird gar nicht erst betrachtet
    "precheck_min_age_days": 365,  # jünger wird nie vorangehakt, höchstens "unklar"
    "delete_at": 60,             # ab dieser Punktzahl vorangehakt
    "unsure_at": 35,             # darunter gilt "behalten"
}

# Obergrenze pro Lauf. Verhindert, dass ein Versehen (falscher Ordner, zu
# scharfe Einstellung) auf einen Schlag das halbe Postfach räumt.
MAX_PER_RUN = 2000

# Absender, die praktisch nie eine Antwort erwarten.
AUTOMATED_PREFIXES = (
    "noreply", "no-reply", "no_reply", "donotreply", "do-not-reply",
    "notifications", "notification", "newsletter", "mailer-daemon",
    "bounce", "bounces", "automail", "auto-confirm", "postmaster",
)


# ---------------------------------------------------------------------------
# Konten
# ---------------------------------------------------------------------------

def load_accounts():
    if not ACCOUNTS_FILE.exists():
        return []
    try:
        data = json.loads(ACCOUNTS_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        raise RuntimeError(f"{ACCOUNTS_FILE} ist beschädigt: {e}") from e
    return data.get("accounts", [])


def save_accounts(accounts):
    ACCOUNTS_FILE.write_text(
        json.dumps({"accounts": accounts}, indent=2, ensure_ascii=False),
        encoding="utf-8")


def account_by_name(name):
    for acc in load_accounts():
        if acc["name"] == name:
            return acc
    raise RuntimeError(f"Kein Konto namens '{name}'. Vorhanden: "
                       f"{', '.join(a['name'] for a in load_accounts()) or '(keins)'}")


def add_account(name, host, user, password, port=993):
    accounts = [a for a in load_accounts() if a["name"] != name]
    accounts.append({"name": name, "host": host, "port": int(port),
                     "user": user, "password": password})
    save_accounts(accounts)
    return accounts


def remove_account(name):
    accounts = [a for a in load_accounts() if a["name"] != name]
    save_accounts(accounts)
    return accounts


def connect_account(account):
    return mi.connect(account["host"], account["user"], account["password"],
                      account.get("port", 993))


def test_account(account):
    """Verbindung prüfen und Ordner zählen - für den 'Testen'-Knopf."""
    conn = connect_account(account)
    try:
        folders = mi.list_folders(conn)
        return {"ok": True, "folders": len(folders),
                "trash": mi.find_trash(folders)}
    finally:
        mi.close(conn)


def diagnose_account(account, progress=None):
    """Postfach durchleuchten - streng nur lesend.

    Beantwortet die Frage "warum sehe ich im Mailprogramm keine Löschungen?"
    ohne Raten: Kann der Server überhaupt verschieben? Welchen Ordner hält das
    Werkzeug für den Papierkorb, und woran hat es das festgemacht? Und liegen
    irgendwo Mails herum, die nur als gelöscht markiert sind?
    """
    conn = connect_account(account)
    try:
        folders = mi.list_folders(conn)
        trash, trash_via = mi.find_trash_detail(folders)
        method = mi.planned_method(conn)
        out = {
            "account": account["name"],
            "can_move": mi.has_capability(conn, "MOVE"),
            "can_uidplus": mi.has_capability(conn, "UIDPLUS"),
            "method": method,
            "method_text": mi.METHOD_DE[method],
            "trash": trash,
            "trash_via": trash_via,
            "folders": [],
        }
        for name, flags in folders:
            if "\\noselect" in [f.lower() for f in flags]:
                continue
            if progress:
                progress(f"{account['name']}: zähle '{name}' ...")
            try:
                total, deleted = mi.folder_stats(conn, name)
                error = None
            except Exception as e:  # noqa: BLE001 - ein sperriger Ordner darf
                total = deleted = None      # den ganzen Bericht nicht kippen
                error = str(e)
            out["folders"].append({
                "name": name, "flags": flags, "total": total,
                "deleted": deleted, "is_trash": name == trash, "error": error,
            })
        return out
    finally:
        mi.close(conn)


def list_account_folders(account):
    """Ordner eines Kontos mit Kennzeichen, ob sie zum Scannen taugen."""
    conn = connect_account(account)
    try:
        out = []
        for name, flags in mi.list_folders(conn):
            if "\\noselect" in [f.lower() for f in flags]:
                continue
            out.append({"name": name, "skip": mi.is_skipped_folder(name, flags)})
        return out
    finally:
        mi.close(conn)


# ---------------------------------------------------------------------------
# Lernspeicher
# ---------------------------------------------------------------------------

def load_decisions():
    if not DECISIONS_FILE.exists():
        return {"senders": {}, "domains": {}}
    try:
        data = json.loads(DECISIONS_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"senders": {}, "domains": {}}
    data.setdefault("senders", {})
    data.setdefault("domains", {})
    return data


def save_decisions(decisions):
    DECISIONS_FILE.write_text(json.dumps(decisions, indent=2, ensure_ascii=False),
                              encoding="utf-8")


def record_decisions(mails, decisions=None):
    """Merkt sich die bestätigten Entscheidungen - in BEIDE Richtungen.

    Abgehakte Mails zählen als "gelöscht", bewusst nicht abgehakte Kandidaten
    als "behalten". Ohne die Behalten-Seite würde das Tool nur lernen, wen es
    wegwerfen soll, aber nie, wen es künftig in Ruhe lassen muss.
    """
    decisions = decisions if decisions is not None else load_decisions()
    today = datetime.now(timezone.utc).date().isoformat()

    for mail in mails:
        sender = (mail.get("from") or "").strip().lower()
        if not sender:
            continue
        domain = mail.get("domain") or sender.rpartition("@")[2]
        field = "deleted" if mail.get("delete") else "kept"
        for store, key in ((decisions["senders"], sender), (decisions["domains"], domain)):
            if not key:
                continue
            entry = store.setdefault(key, {"deleted": 0, "kept": 0, "last": ""})
            entry[field] = entry.get(field, 0) + 1
            entry["last"] = today
    return decisions


def learned_summary(decisions=None, limit=40):
    """Was das Tool bisher gelernt hat - für Anzeige und Nachvollziehbarkeit."""
    decisions = decisions if decisions is not None else load_decisions()
    rows = []
    for sender, e in decisions.get("senders", {}).items():
        deleted, kept = e.get("deleted", 0), e.get("kept", 0)
        rows.append({"sender": sender, "deleted": deleted, "kept": kept,
                     "last": e.get("last", ""),
                     "tendency": "löschen" if deleted > kept else
                                 ("behalten" if kept > deleted else "gemischt")})
    rows.sort(key=lambda r: -(r["deleted"] + r["kept"]))
    return rows[:limit]


# ---------------------------------------------------------------------------
# Eigene Regeln
# ---------------------------------------------------------------------------

def load_rules():
    """Von Hand gepflegte Regeln. Anders als der Lernspeicher sind das
    ausdrückliche Anweisungen - sie schlagen deshalb das Gelernte."""
    rules = {name: [] for name in RULE_LISTS}
    if not RULES_FILE.exists():
        return rules
    try:
        data = json.loads(RULES_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return rules
    for name in RULE_LISTS:
        rules[name] = [str(v).strip() for v in data.get(name, []) if str(v).strip()]
    return rules


def save_rules(rules):
    RULES_FILE.write_text(
        json.dumps({name: rules.get(name, []) for name in RULE_LISTS},
                   indent=2, ensure_ascii=False), encoding="utf-8")


def add_rule(list_name, value):
    if list_name not in RULE_LISTS:
        raise RuntimeError(f"Unbekannte Regelliste '{list_name}'.")
    value = str(value).strip()
    if not value:
        raise RuntimeError("Leerer Regeleintrag.")
    rules = load_rules()
    lowered = [v.lower() for v in rules[list_name]]
    if value.lower() not in lowered:
        rules[list_name].append(value)
        save_rules(rules)
    return rules


def remove_rule(list_name, value):
    if list_name not in RULE_LISTS:
        raise RuntimeError(f"Unbekannte Regelliste '{list_name}'.")
    rules = load_rules()
    rules[list_name] = [v for v in rules[list_name] if v.lower() != str(value).strip().lower()]
    save_rules(rules)
    return rules


def _matches_sender(sender, patterns):
    """Trifft die Absenderadresse eine der Regeln? Erlaubt die genaue Adresse
    oder '*@domain.de' für alles von einer Domain."""
    sender = (sender or "").strip().lower()
    if not sender:
        return None
    domain = sender.rpartition("@")[2]
    for pattern in patterns:
        p = pattern.strip().lower()
        if not p:
            continue
        if p.startswith("*@"):
            if domain == p[2:]:
                return pattern
        elif p == sender:
            return pattern
    return None


def _matches_subject(subject, patterns):
    """Einfache Teiltext-Suche ohne Groß-/Kleinschreibung - bewusst keine
    regulären Ausdrücke, damit sich die Regeln ohne Vorkenntnisse pflegen
    lassen und nicht versehentlich zu viel treffen."""
    text = (subject or "").lower()
    if not text:
        return None
    for pattern in patterns:
        p = pattern.strip().lower()
        if p and p in text:
            return pattern
    return None


# ---------------------------------------------------------------------------
# Bewertung
# ---------------------------------------------------------------------------

def _age_days(mail, now=None):
    raw = mail.get("date")
    if not raw:
        return None
    try:
        dt = datetime.fromisoformat(raw)
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    now = now or datetime.now(timezone.utc)
    return max(0, (now - dt).days)


def _is_bulk(mail):
    if mail.get("list_unsubscribe") or mail.get("list_id"):
        return True
    if mail.get("precedence") in ("bulk", "list", "junk"):
        return True
    auto = mail.get("auto_submitted", "")
    return bool(auto) and auto != "no"


def _is_automated_sender(mail):
    local = (mail.get("from") or "").partition("@")[0].lower()
    return any(local.startswith(p) for p in AUTOMATED_PREFIXES)


def _learned_hint(mail, decisions):
    """Was frühere Entscheidungen über diesen Absender sagen.

    Gibt (punkte, begründung, urteil) zurück. urteil ist None, "keep" oder
    "delete" - bei eindeutiger Vorgeschichte entscheidet das Gelernte direkt,
    statt nur Punkte beizusteuern. Wer dreimal dasselbe weggeworfen und nie
    etwas davon behalten hat, hat seine Meinung deutlich genug gesagt; ein
    Punktesystem würde dieses klare Signal nur verwässern.
    """
    sender = (mail.get("from") or "").strip().lower()
    entry = decisions.get("senders", {}).get(sender)
    if entry:
        deleted, kept = entry.get("deleted", 0), entry.get("kept", 0)
        if kept >= 2 and deleted == 0:
            return 0, f"von dir bisher immer behalten ({kept}×)", "keep"
        if deleted >= 3 and kept == 0:
            return 0, f"von dir bisher immer gelöscht ({deleted}×)", "delete"
        if deleted >= 1 and kept == 0:
            return 15, f"von dir schon gelöscht ({deleted}×)", None
        return 0, "", None

    # Nur wenn die genaue Adresse unbekannt ist, zählt die Domain - und
    # schwächer, weil hinter einer Domain auch persönliche Post stecken kann.
    domain = mail.get("domain") or sender.rpartition("@")[2]
    entry = decisions.get("domains", {}).get(domain)
    if entry:
        deleted, kept = entry.get("deleted", 0), entry.get("kept", 0)
        if deleted >= 5 and kept == 0:
            return 15, f"Absender {domain} bisher immer gelöscht ({deleted}×)", None
    return 0, "", None


def score_mail(mail, decisions=None, settings=None, now=None, rules=None):
    """Bewertet eine Mail: (punkte, empfehlung, begründungen).

    empfehlung ist "delete", "unsure" oder "keep". Nur "delete" wird in Excel
    und Web-Oberfläche vorangehakt - alles andere muss der Nutzer selbst
    ankreuzen.

    Reihenfolge der Entscheidung (die erste greifende gewinnt):
      1. eigene "nie löschen"-Regeln
      2. harte Schutzregeln (markiert, beantwortet, Entwurf, zu neu)
      3. eigene "immer löschen"-Regeln
      4. Gelerntes aus früheren Entscheidungen
      5. Punktebewertung
    Ausdrückliche Anweisungen des Nutzers stehen also über dem Gelernten,
    aber unter den Schutzregeln - eine markierte oder ganz frische Mail wird
    auch dann nicht vorgeschlagen, wenn ihr Absender auf "immer löschen" steht.
    """
    decisions = decisions if decisions is not None else load_decisions()
    rules = rules if rules is not None else load_rules()
    cfg = {**DEFAULT_SETTINGS, **(settings or {})}
    flags = [f.lower() for f in mail.get("flags", [])]
    reasons = []

    # --- 1. Eigene "nie löschen"-Regeln ---
    hit = _matches_sender(mail.get("from"), rules.get("never_delete", []))
    if hit:
        return 0, "keep", [f"eigene Regel: nie löschen ({hit})"]
    hit = _matches_subject(mail.get("subject"), rules.get("subject_never", []))
    if hit:
        return 0, "keep", [f"eigene Betreffregel: nie löschen („{hit}“)"]

    # --- Harte Schutzregeln: nie vorschlagen, egal wie die Punkte stehen ---
    if "\\flagged" in flags:
        return 0, "keep", ["markiert - wird nie vorgeschlagen"]
    if "\\answered" in flags:
        return 0, "keep", ["du hast geantwortet - wird nie vorgeschlagen"]
    if "\\draft" in flags:
        return 0, "keep", ["Entwurf - wird nie vorgeschlagen"]

    age = _age_days(mail, now)
    if age is None:
        return 0, "keep", ["kein lesbares Datum - sicherheitshalber behalten"]
    if age < cfg["min_age_days"]:
        return 0, "keep", [f"erst {age} Tage alt - wird nie vorgeschlagen"]

    # --- 3. Eigene "immer löschen"-Regeln (nach den Schutzregeln!) ---
    hit = _matches_sender(mail.get("from"), rules.get("always_delete", []))
    if hit:
        return cfg["delete_at"], "delete", [f"eigene Regel: immer löschen ({hit})"]

    # --- 4. Gelerntes ---
    learned_points, learned_reason, learned_verdict = _learned_hint(mail, decisions)
    if learned_verdict == "keep":
        return 0, "keep", [learned_reason + " - wird nie vorgeschlagen"]
    if learned_verdict == "delete":
        return cfg["delete_at"], "delete", [learned_reason]

    # --- Punktevergabe ---
    score = 0

    for threshold, points, label in ((1825, 30, "über 5 Jahre alt"),
                                     (1095, 25, "über 3 Jahre alt"),
                                     (730, 20, "über 2 Jahre alt"),
                                     (365, 15, "über 1 Jahr alt"),
                                     (180, 10, "über ein halbes Jahr alt"),
                                     (90, 5, "über 3 Monate alt")):
        if age >= threshold:
            score += points
            reasons.append(label)
            break

    if _is_bulk(mail):
        score += 35
        reasons.append("Newsletter/Massenmail")
    if _is_automated_sender(mail):
        score += 30
        reasons.append("automatischer Absender")

    hit = _matches_subject(mail.get("subject"), rules.get("subject_delete", []))
    if hit:
        # Starkes Signal, aber kein Machtwort: Betreffzeilen treffen gröber
        # als Adressen. Zusammen mit dem Alter reicht es für einen Vorschlag,
        # auf frischer Post feuert es dank Voranhak-Sperre nicht.
        score += 35
        reasons.append(f"eigene Betreffregel („{hit}“)")

    if "\\seen" in flags:
        score += 10
        reasons.append("gelesen")
    else:
        score += 5
        reasons.append("nie gelesen")

    size = mail.get("size", 0)
    if size >= 5_000_000:
        score += 10
        reasons.append(f"sehr groß ({size // 1_000_000} MB)")
    elif size >= 1_000_000:
        score += 5
        reasons.append(f"groß ({size // 1_000_000} MB)")

    if learned_points:
        score += learned_points
        reasons.append(learned_reason)

    score = max(0, min(100, score))
    if score >= cfg["delete_at"]:
        recommendation = "delete"
    elif score >= cfg["unsure_at"]:
        recommendation = "unsure"
    else:
        recommendation = "keep"

    # Voranhak-Sperre: unterhalb dieser Grenze wird nie von selbst angehakt,
    # höchstens "unklar" vorgeschlagen. Ohne diese Grenze addieren sich
    # Einzelsignale zu einem Vorschlag, obwohl die Mail noch recht frisch ist -
    # ein Newsletter von "newsletter@..." käme sonst schon nach wenigen Wochen
    # über die Schwelle, nur weil "Newsletter" und "automatischer Absender"
    # zusammenfallen.
    if recommendation == "delete" and age < cfg["precheck_min_age_days"]:
        recommendation = "unsure"
        reasons.append(f"noch keine {cfg['precheck_min_age_days'] // 365 or 1} Jahr(e) alt "
                       "- nicht von selbst angehakt")
    return score, recommendation, reasons


# ---------------------------------------------------------------------------
# Scannen
# ---------------------------------------------------------------------------

RECOMMENDATION_DE = {"delete": "löschen", "unsure": "unklar", "keep": "behalten"}


def scan_accounts(selection, settings=None, progress=None):
    """Durchsucht die gewählten Konten/Ordner - streng nur lesend.

    selection: [{"account": "iCloud", "folders": ["INBOX", "Archiv"]}, ...]
    Gibt die Scan-Struktur zurück (und speichert sie über save_scan()).
    """
    cfg = {**DEFAULT_SETTINGS, **(settings or {})}
    decisions = load_decisions()
    rules = load_rules()   # einmal laden statt pro Mail
    scan = {"created": datetime.now(timezone.utc).isoformat(),
            "settings": cfg, "accounts": []}

    for entry in selection:
        account = account_by_name(entry["account"])
        conn = connect_account(account)
        try:
            folders = mi.list_folders(conn)
            trash = mi.find_trash(folders)
            acc_out = {"account": account["name"], "trash": trash, "folders": []}

            for folder in entry.get("folders", []):
                if progress:
                    progress(f"{account['name']}: lese '{folder}' ...")

                def on_block(done, total, _f=folder, _a=account["name"]):
                    if progress:
                        progress(f"{_a}: '{_f}' {done}/{total} Mails gelesen ...")

                mails, uidvalidity = mi.scan_folder(conn, folder, progress=on_block)
                for mail in mails:
                    score, recommendation, reasons = score_mail(mail, decisions, cfg,
                                                                rules=rules)
                    mail["score"] = score
                    mail["recommendation"] = recommendation
                    mail["reasons"] = reasons
                    mail["delete"] = recommendation == "delete"
                acc_out["folders"].append(
                    {"folder": folder, "uidvalidity": uidvalidity, "mails": mails})
            scan["accounts"].append(acc_out)
        finally:
            mi.close(conn)

    save_scan(scan)
    return scan


def save_scan(scan):
    SCAN_FILE.parent.mkdir(exist_ok=True)
    SCAN_FILE.write_text(json.dumps(scan, ensure_ascii=False), encoding="utf-8")


def load_scan():
    if not SCAN_FILE.exists():
        raise RuntimeError("Es liegt kein Scan vor. Bitte zuerst 'scan' ausführen.")
    return json.loads(SCAN_FILE.read_text(encoding="utf-8"))


def clear_scan():
    """Gespeicherten Scan verwerfen.

    Die Liste ist eine Momentaufnahme. Wenn sie nicht mehr zum Postfach passt,
    ist sie wertlos - dann lieber wegwerfen und neu einlesen, als weiter mit
    einem Stand arbeiten, dem niemand trauen kann.
    """
    existed = SCAN_FILE.exists()
    SCAN_FILE.unlink(missing_ok=True)
    return existed


def iter_mails(scan):
    """Alle Mails eines Scans als (konto, ordner_eintrag, mail)."""
    for acc in scan.get("accounts", []):
        for folder in acc.get("folders", []):
            for mail in folder.get("mails", []):
                yield acc, folder, mail


def mail_key(account_name, folder_name, uid):
    return f"{account_name}|{folder_name}|{uid}"


def scan_summary(scan):
    counts = {"total": 0, "delete": 0, "unsure": 0, "keep": 0,
              "bytes": 0, "selected": 0, "scanned_bytes": 0, "attachments": 0}
    for _acc, _folder, mail in iter_mails(scan):
        counts["total"] += 1
        counts[mail.get("recommendation", "keep")] += 1
        counts["scanned_bytes"] += mail.get("size", 0)
        if mail.get("has_attachment"):
            counts["attachments"] += 1
        if mail.get("delete"):
            counts["selected"] += 1
            counts["bytes"] += mail.get("size", 0)   # was die Auswahl freigibt
    return counts


# ---------------------------------------------------------------------------
# Excel-Rundlauf
# ---------------------------------------------------------------------------

EXCEL_COLUMNS = [
    "Konto", "Ordner", "UID", "Löschen", "Empfehlung", "Punkte", "Begründung",
    "Datum", "Alter (Tage)", "Von", "Absender", "Domain", "Betreff",
    "Größe (KB)", "Anhang", "Anhangnamen", "Gelesen", "Markiert", "Beantwortet",
]
_EXCEL_WIDTHS = [16, 22, 10, 9, 12, 8, 46, 12, 12, 26, 34, 22, 52,
                 12, 9, 32, 9, 9, 12]


def scan_to_excel(scan, path):
    """Scan-Ergebnis als Excel-Liste - Aufbau wie bei den Kontakten, damit der
    Ablauf vertraut bleibt (Kopfzeile fixiert, Filter an, Löschen-Spalte vorn)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Aufräumen"
    ws.append(EXCEL_COLUMNS)
    fill = PatternFill("solid", fgColor="305496")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(EXCEL_COLUMNS) + 1):
        cell = ws.cell(1, col)
        cell.fill, cell.font = fill, font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    rows = 0
    for acc, folder, mail in iter_mails(scan):
        flags = [f.lower() for f in mail.get("flags", [])]
        age = _age_days(mail)
        ws.append([
            acc["account"], folder["folder"], mail["uid"],
            "x" if mail.get("delete") else "",
            RECOMMENDATION_DE.get(mail.get("recommendation"), ""),
            mail.get("score", 0), ", ".join(mail.get("reasons", [])),
            (mail.get("date") or "")[:10], age if age is not None else "",
            mail.get("from_name", ""), mail.get("from", ""), mail.get("domain", ""),
            mail.get("subject", ""), round(mail.get("size", 0) / 1024),
            "ja" if mail.get("has_attachment") else "",
            ", ".join(mail.get("attachments", [])),
            "ja" if "\\seen" in flags else "",
            "ja" if "\\flagged" in flags else "",
            "ja" if "\\answered" in flags else "",
        ])
        rows += 1

    ws.freeze_panes = "E2"   # Konto/Ordner/UID/Löschen bleiben sichtbar
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{rows + 1}"
    for col, width in enumerate(_EXCEL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.column_dimensions["C"].hidden = True   # UID nicht versehentlich ändern
    wb.save(path)
    return rows


def scan_from_excel(scan, path):
    """Liest die bearbeitete Excel-Liste zurück und überträgt die Häkchen der
    Spalte 'Löschen' in den Scan. Nutzt dieselbe Auswertung wie bei den
    Kontakten (_cell_is_delete), damit 'x', 'ja', '1' ... gleich wirken."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c or "").strip() for c in next(rows, [])]
    for needed in ("Konto", "Ordner", "UID", "Löschen"):
        if needed not in header:
            raise RuntimeError(
                f"Spalte '{needed}' fehlt. Stammt die Datei aus "
                "'Als Excel herunterladen' bzw. 'to-excel'?")
    idx = {name: header.index(name) for name in ("Konto", "Ordner", "UID", "Löschen")}

    wanted = {}
    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        try:
            uid = int(row[idx["UID"]])
        except (TypeError, ValueError):
            continue
        key = mail_key(str(row[idx["Konto"]]), str(row[idx["Ordner"]]), uid)
        wanted[key] = _cell_is_delete(row[idx["Löschen"]])
    wb.close()

    changed = 0
    for acc, folder, mail in iter_mails(scan):
        key = mail_key(acc["account"], folder["folder"], mail["uid"])
        if key in wanted and bool(mail.get("delete")) != wanted[key]:
            mail["delete"] = wanted[key]
            changed += 1
    save_scan(scan)
    return {"gelesen": len(wanted), "geändert": changed,
            "ausgewählt": sum(1 for _a, _f, m in iter_mails(scan) if m.get("delete"))}


def apply_selection(scan, keys):
    """Auswahl aus der Web-Oberfläche übernehmen (Liste von mail_key-Strings)."""
    wanted = set(keys)
    for acc, folder, mail in iter_mails(scan):
        mail["delete"] = mail_key(acc["account"], folder["folder"], mail["uid"]) in wanted
    save_scan(scan)
    return sum(1 for _a, _f, m in iter_mails(scan) if m.get("delete"))


# ---------------------------------------------------------------------------
# Ausführen
# ---------------------------------------------------------------------------

def clean(scan, dry_run=True, force=False, progress=None):
    """Verschiebt die angehakten Mails in den Papierkorb des jeweiligen Kontos.

    Testlauf ist der Standard. Beim echten Lauf werden anschließend die
    Entscheidungen gelernt - und zwar nur über Mails, die überhaupt zur
    Auswahl standen (Empfehlung 'löschen' oder 'unklar'). Über Mails, die nie
    vorgeschlagen wurden, hat der Nutzer auch nichts entschieden; sie als
    'behalten' zu zählen würde den Lernspeicher mit Zufallsdaten fluten.
    """
    result = {"moved": 0, "flagged": 0, "failed": 0, "skipped": 0, "errors": 0,
              "total": 0, "trash": {}, "log": [], "dry_run": dry_run}

    selected = [(acc, folder, mail) for acc, folder, mail in iter_mails(scan)
                if mail.get("delete")]
    result["total"] = len(selected)
    if not selected:
        result["log"].append("Nichts ausgewählt - es gibt nichts zu tun.")
        return result

    if len(selected) > MAX_PER_RUN and not force:
        raise RuntimeError(
            f"{len(selected)} Mails ausgewählt - das ist mehr als die "
            f"Sicherheitsgrenze von {MAX_PER_RUN} pro Lauf. Bitte die Auswahl "
            "prüfen (stimmen Ordner und Einstellungen?) oder den Lauf mit "
            "--force bzw. in kleineren Portionen wiederholen.")

    # Vor dem Lauf festhalten, woraus gelernt wird: weiter unten fallen die
    # erledigten Mails aus dem Scan heraus, und ausgerechnet die gelöschten sind
    # das wichtigste Signal. Die Dicts bleiben gültig, auch wenn die Liste, in
    # der sie stehen, später ausgedünnt wird.
    reviewed = [m for _a, _f, m in iter_mails(scan)
                if m.get("recommendation") in ("delete", "unsure")]

    for acc in scan.get("accounts", []):
        by_folder = {}
        for folder in acc.get("folders", []):
            items = [{"uid": m["uid"], "message_id": m.get("message_id", "")}
                     for m in folder.get("mails", []) if m.get("delete")]
            if items:
                by_folder[folder["folder"]] = (items, folder.get("uidvalidity"))
        if not by_folder:
            continue

        trash = acc.get("trash")
        if not trash:
            count = sum(len(v[0]) for v in by_folder.values())
            result["errors"] += count
            result["log"].append(
                f"FEHLER: Für Konto '{acc['account']}' wurde kein Papierkorb "
                "gefunden - nichts verschoben.")
            continue

        try:
            account = account_by_name(acc["account"])
            conn = connect_account(account)
        except Exception as e:  # noqa: BLE001
            count = sum(len(v[0]) for v in by_folder.values())
            result["errors"] += count
            result["log"].append(f"FEHLER: Konto '{acc['account']}' nicht erreichbar: {e}")
            continue

        result["trash"][acc["account"]] = trash
        try:
            for folder_name, (items, uidvalidity) in by_folder.items():
                if progress:
                    progress(f"{acc['account']}: '{folder_name}' - {len(items)} Mails ...")
                part = mi.move_to_trash(conn, folder_name, items, trash,
                                        expected_uidvalidity=uidvalidity,
                                        dry_run=dry_run)
                for key in ("moved", "flagged", "failed", "skipped", "errors"):
                    result[key] += part[key]
                # Immer sagen, wohin - nicht nur im Fehlerfall. Ohne diese Zeile
                # steht im Protokoll eines erfolgreichen Laufs gar nichts, und
                # niemand kann nachvollziehen, wo die Mails gelandet sind.
                if not dry_run and part["moved"]:
                    result["log"].append(
                        f"{acc['account']} / {folder_name}: {part['moved']} Mails nach "
                        f"'{trash}' verschoben (nachgeprüft: sie sind dort weg).")
                result["log"].extend(
                    f"{acc['account']} / {folder_name}: {line}" for line in part["log"])
                if not dry_run:
                    _mark_verified(acc, folder_name, part)
                    _forget_moved(scan, acc, folder_name, part)
        finally:
            mi.close(conn)

    if not dry_run:
        # "delete" ist nur die Auswahl - ob die Mail wirklich weg ist, weiß
        # allein die Nachprüfung in move_to_trash(). Nur bewusst behaltene
        # Mails (delete=False) lernen ungeprüft, weil der Server sie nie
        # angefasst hat; bewusst gelöschte nur, wenn verified_deleted gesetzt
        # wurde. Sonst lernt das Tool aus Mails, die noch im Postfach liegen -
        # genau der gemeldete Fehler.
        to_learn = [m for m in reviewed
                   if not m.get("delete") or m.get("verified_deleted")]
        save_decisions(record_decisions(to_learn))
        note = (f"Gelernt aus {len(to_learn)} geprüften Vorschlägen "
                "(gelöscht wie behalten).")
        open_count = len(reviewed) - len(to_learn)
        if open_count:
            note += (f" {open_count} Auswahlen sind noch offen (nicht bestätigt "
                     "verschoben) und wurden nicht gelernt.")
        result["log"].append(note)
        scan["executed"] = datetime.now(timezone.utc).isoformat()
        save_scan(scan)
    return result


def _mark_verified(acc, folder_name, part):
    """Nachweislich verschobene Mail-Dicts kennzeichnen - fürs Lernen unten.

    Läuft vor _forget_moved(), das dieselben Mails gleich aus dem Scan nimmt;
    die Markierung bleibt trotzdem gültig, weil reviewed() dieselben Objekte
    referenziert.
    """
    gone = set(part.get("moved_uids") or ())
    if not gone:
        return
    for folder in acc.get("folders", []):
        if folder["folder"] == folder_name:
            for mail in folder.get("mails", []):
                if mail["uid"] in gone:
                    mail["verified_deleted"] = True


def _forget_moved(scan, acc, folder_name, part):
    """Erledigte Mails aus dem Scan nehmen - und nur die.

    Entfernt wird ausschließlich, was die Nachprüfung im Postfach nicht mehr
    gefunden hat. Angehakt sein reicht nicht: übersprungene Mails (hinter der
    Nummer steckt inzwischen eine andere) und nur markierte liegen weiter im
    Postfach und müssen deshalb auch in der Liste stehen bleiben. Sonst sieht
    die Liste sauberer aus als das Postfach - und genau das ist unbrauchbar.
    """
    gone = set(part.get("moved_uids") or ())
    if not gone:
        return
    for folder in acc.get("folders", []):
        if folder["folder"] == folder_name:
            folder["mails"] = [m for m in folder.get("mails", [])
                               if m["uid"] not in gone]


# ---------------------------------------------------------------------------
# Kommandozeile
# ---------------------------------------------------------------------------

def _print_accounts():
    accounts = load_accounts()
    if not accounts:
        print("Noch keine Konten eingerichtet. Anlegen mit:")
        print("  python mail_cleanup.py konten --add")
        return
    print(f"{len(accounts)} Konto/Konten:")
    for acc in accounts:
        print(f"  - {acc['name']}: {acc['user']} auf {acc['host']}:{acc.get('port', 993)}")


def cmd_konten(args):
    if args.add:
        print("Neues Konto anlegen (Abbruch mit Strg+C)\n")
        name = input("Anzeigename (z.B. iCloud): ").strip()
        print(f"Anbieter-Kürzel für die Voreinstellung: {', '.join(PRESETS)}")
        provider = input("Anbieter (oder leer für eigenen Server): ").strip().lower()
        preset = PRESETS.get(provider, {})
        host = input(f"IMAP-Server [{preset.get('host', '')}]: ").strip() or preset.get("host", "")
        port = input(f"Port [{preset.get('port', 993)}]: ").strip() or preset.get("port", 993)
        user = input("Benutzername (meist die E-Mail-Adresse): ").strip()
        print("Bei iCloud und Gmail wird ein APP-SPEZIFISCHES Passwort gebraucht,")
        print("nicht das normale Kennwort.")
        # Ohne diesen Hinweis wirkt die Eingabe wie eingefroren: getpass zeigt
        # beim Tippen bewusst gar nichts an - keine Zeichen, keine Sternchen.
        print("Hinweis: Die Eingabe bleibt unsichtbar (auch keine Sternchen).")
        print("         Einfach tippen bzw. einfügen und Enter drücken.")
        password = getpass("Passwort: ")
        if not (name and host and user and password):
            print("Abgebrochen - es fehlten Angaben.")
            sys.exit(1)
        add_account(name, host, user, password, port)
        print(f"\nKonto '{name}' gespeichert in {ACCOUNTS_FILE}.")
        try:
            info = test_account(account_by_name(name))
            print(f"Verbindung ok - {info['folders']} Ordner, "
                  f"Papierkorb: {info['trash'] or 'nicht gefunden!'}")
        except Exception as e:  # noqa: BLE001
            print(f"ACHTUNG: Verbindung fehlgeschlagen: {e}")
        return

    if args.remove:
        remove_account(args.remove)
        print(f"Konto '{args.remove}' entfernt.")
        return

    if args.test:
        for acc in load_accounts():
            try:
                info = test_account(acc)
                print(f"  {acc['name']}: ok, {info['folders']} Ordner, "
                      f"Papierkorb: {info['trash'] or 'NICHT GEFUNDEN'}")
            except Exception as e:  # noqa: BLE001
                print(f"  {acc['name']}: FEHLER - {e}")
        return

    if args.folders:
        acc = account_by_name(args.folders)
        for entry in list_account_folders(acc):
            mark = "  (wird übersprungen)" if entry["skip"] else ""
            print(f"  {entry['name']}{mark}")
        return

    _print_accounts()


def cmd_scan(args):
    if args.reset:
        if clear_scan():
            print(f"Gespeicherter Scan verworfen ({SCAN_FILE}).")
        else:
            print("Es lag kein Scan vor - nichts zu verwerfen.")
        print("Neu einlesen mit:  python mail_cleanup.py scan")
        return

    accounts = load_accounts()
    if not accounts:
        print("Noch keine Konten eingerichtet: python mail_cleanup.py konten --add")
        sys.exit(1)

    names = args.account or [a["name"] for a in accounts]
    selection = []
    for name in names:
        acc = account_by_name(name)
        if args.folder:
            folders = list(args.folder)
        else:
            folders = [f["name"] for f in list_account_folders(acc) if not f["skip"]]
            print(f"{name}: {len(folders)} Ordner werden gelesen "
                  f"(Papierkorb, Entwürfe und Spam bleiben außen vor).")
        selection.append({"account": name, "folders": folders})

    settings = {"min_age_days": args.min_age} if args.min_age is not None else None
    scan = scan_accounts(selection, settings, progress=lambda m: print(f"  {m}", flush=True))

    counts = scan_summary(scan)
    print(f"\n{counts['total']} Mails geprüft: {counts['delete']} zum Löschen "
          f"vorgeschlagen, {counts['unsure']} unklar, {counts['keep']} behalten.")
    print(f"Vorgeschlagen wären {counts['bytes'] / 1_048_576:.1f} MB.")
    print(f"\nGespeichert: {SCAN_FILE}")
    print("Weiter mit:  python mail_cleanup.py to-excel")


def cmd_to_excel(args):
    rows = scan_to_excel(load_scan(), args.output)
    print(f"{rows} Mails geschrieben.")
    print(f"Gespeichert: {args.output}")
    print("Jetzt in Excel die Spalte 'Löschen' prüfen, dann:")
    print(f"  python mail_cleanup.py from-excel --input {args.output}")


def cmd_from_excel(args):
    info = scan_from_excel(load_scan(), args.input)
    print(f"{info['gelesen']} Zeilen gelesen, {info['geändert']} Häkchen geändert.")
    print(f"{info['ausgewählt']} Mails sind jetzt zum Löschen ausgewählt.")
    print("Weiter mit:  python mail_cleanup.py clean        (Testlauf)")


def cmd_clean(args):
    scan = load_scan()
    dry_run = not args.execute
    if dry_run:
        print("TESTLAUF: Es wird nichts verschoben. Echter Lauf mit --execute.\n")
    result = clean(scan, dry_run=dry_run, force=args.force,
                   progress=lambda m: print(f"  {m}", flush=True))
    for line in result["log"]:
        print(f"  {line}")
    verb = "würden verschoben" if dry_run else "nachweislich verschoben"
    print(f"\n{result['moved']} Mails {verb}, {result['skipped']} übersprungen, "
          f"{result['errors']} Fehler (von {result['total']} ausgewählten).")
    if result["flagged"]:
        print(f"{result['flagged']} Mails wurden NUR ALS GELÖSCHT MARKIERT und "
              "liegen weiter im Ordner.")
    if result["failed"]:
        print(f"{result['failed']} Mails sind trotz Erfolgsmeldung des Servers "
              "noch da.")
    if dry_run and result["moved"]:
        print("\nSieht das gut aus? Dann:")
        print("  python mail_cleanup.py clean --execute")
    if (result["flagged"] or result["failed"]) and not dry_run:
        print("\nWas dahintersteckt, zeigt:")
        print("  python mail_cleanup.py diagnose")


def cmd_diagnose(args):
    """Postfach durchleuchten - beantwortet 'warum sehe ich keine Löschungen?'."""
    accounts = load_accounts()
    if args.account:
        accounts = [account_by_name(args.account)]
    if not accounts:
        print("Noch keine Konten eingerichtet.")
        return
    for account in accounts:
        info = diagnose_account(account, progress=lambda m: print(f"  {m}", flush=True))
        print(f"\n=== {info['account']} ===")
        print(f"Server kann MOVE:    {'ja' if info['can_move'] else 'NEIN'}")
        print(f"Server kann UIDPLUS: {'ja' if info['can_uidplus'] else 'NEIN'}")
        print(f"Verfahren beim Aufräumen: {info['method_text']}")
        if info["method"] == "copy_flag":
            print("  -> Dieser Server verschiebt NICHT. Mails werden in den")
            print("     Papierkorb kopiert und im Ordner nur markiert. Genau")
            print("     deshalb sieht man im Mailprogramm keine Löschung.")
        via = {"flag": "über das \\Trash-Kennzeichen des Servers",
               "name": "ÜBER DEN NAMEN GERATEN", None: ""}[info["trash_via"]]
        print(f"Papierkorb: {info['trash'] or 'NICHT GEFUNDEN'} {via}".rstrip())
        print(f"\n{'Ordner':<34} {'Mails':>8} {'davon gelöscht-markiert':>24}")
        for folder in info["folders"]:
            mark = " <- Papierkorb" if folder["is_trash"] else ""
            if folder["error"]:
                print(f"{folder['name']:<34} {'?':>8} {folder['error']}")
                continue
            print(f"{folder['name']:<34} {folder['total']:>8} "
                  f"{folder['deleted']:>24}{mark}")


_RULE_LABELS = {
    "never_delete": "nie löschen (Absender)",
    "always_delete": "immer löschen (Absender)",
    "subject_never": "nie löschen (Betreff enthält)",
    "subject_delete": "eher löschen (Betreff enthält)",
}


def cmd_regeln(args):
    changes = [("never_delete", args.nie), ("always_delete", args.immer),
               ("subject_never", args.betreff_nie), ("subject_delete", args.betreff_immer)]
    touched = False
    for list_name, values in changes:
        for value in values or []:
            add_rule(list_name, value)
            print(f"Ergänzt in '{_RULE_LABELS[list_name]}': {value}")
            touched = True
    for value in args.remove or []:
        for list_name in RULE_LISTS:
            remove_rule(list_name, value)
        print(f"Entfernt (aus allen Listen): {value}")
        touched = True
    if touched:
        print()

    rules = load_rules()
    if not any(rules[name] for name in RULE_LISTS):
        print("Noch keine eigenen Regeln. Beispiele:")
        print("  python mail_cleanup.py regeln --immer 'newsletter@shop.de'")
        print("  python mail_cleanup.py regeln --immer '*@werbung.de'")
        print("  python mail_cleanup.py regeln --nie 'chef@firma.de'")
        print("  python mail_cleanup.py regeln --betreff-immer 'Ihre Bestellung'")
        return
    for list_name in RULE_LISTS:
        if rules[list_name]:
            print(f"{_RULE_LABELS[list_name]}:")
            for value in rules[list_name]:
                print(f"  - {value}")


def cmd_gelernt(args):
    if args.reset:
        if DECISIONS_FILE.exists():
            DECISIONS_FILE.unlink()
        print("Lernspeicher zurückgesetzt.")
        return
    rows = learned_summary()
    if not rows:
        print("Noch nichts gelernt - das passiert nach dem ersten echten Aufräumen.")
        return
    print(f"{'Absender':44} {'gelöscht':>9} {'behalten':>9}  Tendenz")
    print("-" * 78)
    for row in rows:
        print(f"{row['sender'][:44]:44} {row['deleted']:>9} {row['kept']:>9}  {row['tendency']}")


def main():
    parser = argparse.ArgumentParser(
        description="E-Mail-Postfächer aufräumen: bewerten, prüfen, in den Papierkorb verschieben")
    sub = parser.add_subparsers(dest="cmd", required=True)

    konten = sub.add_parser("konten", help="IMAP-Konten anzeigen, anlegen, prüfen")
    konten.add_argument("--add", action="store_true", help="Neues Konto anlegen (fragt nach)")
    konten.add_argument("--remove", metavar="NAME", help="Konto entfernen")
    konten.add_argument("--test", action="store_true", help="Alle Konten auf Erreichbarkeit prüfen")
    konten.add_argument("--folders", metavar="NAME", help="Ordner eines Kontos auflisten")

    scan = sub.add_parser("scan", help="Postfächer durchsuchen und bewerten (nur lesend)")
    scan.add_argument("--account", action="append",
                      help="Konto (mehrfach angebbar; Standard: alle)")
    scan.add_argument("--folder", action="append",
                      help="Ordner (mehrfach angebbar; Standard: alle sinnvollen)")
    scan.add_argument("--min-age", type=int, metavar="TAGE",
                      help=f"Mails jünger als N Tage nie vorschlagen "
                           f"(Standard: {DEFAULT_SETTINGS['min_age_days']})")
    scan.add_argument("--reset", action="store_true",
                      help="Gespeicherten Scan verwerfen, ohne neu zu durchsuchen")

    toexcel = sub.add_parser("to-excel", help="Scan-Ergebnis als Excel-Liste ausgeben")
    toexcel.add_argument("--output", default="mail_aufraeumen.xlsx")

    fromexcel = sub.add_parser("from-excel", help="Bearbeitete Excel-Liste einlesen")
    fromexcel.add_argument("--input", required=True)

    clean_p = sub.add_parser("clean", help="Ausgewählte Mails in den Papierkorb verschieben")
    clean_p.add_argument("--execute", action="store_true",
                         help="Wirklich verschieben (ohne diese Angabe nur Testlauf)")
    clean_p.add_argument("--force", action="store_true",
                         help=f"Sicherheitsgrenze von {MAX_PER_RUN} Mails pro Lauf aufheben")

    regeln = sub.add_parser("regeln", help="Eigene Regeln anzeigen und pflegen")
    regeln.add_argument("--nie", action="append", metavar="ABSENDER",
                        help="Nie löschen, z.B. 'chef@firma.de' oder '*@firma.de'")
    regeln.add_argument("--immer", action="append", metavar="ABSENDER",
                        help="Immer vorschlagen, z.B. 'news@shop.de' oder '*@werbung.de'")
    regeln.add_argument("--betreff-nie", action="append", metavar="TEXT",
                        dest="betreff_nie", help="Betreff enthält … -> nie löschen")
    regeln.add_argument("--betreff-immer", action="append", metavar="TEXT",
                        dest="betreff_immer", help="Betreff enthält … -> eher löschen")
    regeln.add_argument("--remove", action="append", metavar="WERT",
                        help="Eintrag aus allen Listen entfernen")

    gelernt = sub.add_parser("gelernt", help="Zeigen, was aus deinen Entscheidungen gelernt wurde")
    gelernt.add_argument("--reset", action="store_true", help="Lernspeicher leeren")

    diagnose = sub.add_parser(
        "diagnose", help="Postfach prüfen: kann der Server verschieben, wo ist "
                         "der Papierkorb? (nur lesend)")
    diagnose.add_argument("--account", metavar="NAME", help="Nur dieses Konto")

    args = parser.parse_args()
    handlers = {"konten": cmd_konten, "scan": cmd_scan, "to-excel": cmd_to_excel,
                "from-excel": cmd_from_excel, "clean": cmd_clean, "regeln": cmd_regeln,
                "gelernt": cmd_gelernt, "diagnose": cmd_diagnose}
    try:
        handlers[args.cmd](args)
    except RuntimeError as e:
        print(f"\nFehler: {e}")
        sys.exit(1)
    except KeyboardInterrupt:
        print("\nAbgebrochen.")
        sys.exit(1)


if __name__ == "__main__":
    main()
