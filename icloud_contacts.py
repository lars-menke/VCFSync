"""
iCloud Kontakte - CardDAV Export & Import
==========================================
Verwendung:
  Export:  python icloud_contacts.py export --output meine_kontakte.vcf
  Import:  python icloud_contacts.py import --input bearbeitete_kontakte.vcf

Zugangsdaten werden beim ersten Start abgefragt und in .env gespeichert,
oder als Umgebungsvariablen gesetzt:
  ICLOUD_USER=lars@icloud.com
  ICLOUD_PASS=xxxx-xxxx-xxxx-xxxx   (app-spezifisches Passwort)
"""

import argparse
import base64
import math
import os
import re
import sys
import time
import uuid
import xml.etree.ElementTree as ET
from getpass import getpass
from pathlib import Path
from urllib.parse import urljoin

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from requests.auth import HTTPBasicAuth

# ---------------------------------------------------------------------------
# Konstanten
# ---------------------------------------------------------------------------
CARDDAV_BASE = "https://contacts.icloud.com"
ENV_FILE = Path(".env")
NS = {
    "d": "DAV:",
    "card": "urn:ietf:params:xml:ns:carddav",
    "cs": "http://calendarserver.org/ns/",
}

# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def load_env():
    """Lädt Zugangsdaten aus .env-Datei (KEY=VALUE Zeilen)."""
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text().splitlines():
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())


def get_credentials():
    """Gibt (user, password) zurück, fragt interaktiv falls nötig."""
    load_env()
    user = os.environ.get("ICLOUD_USER") or input("Apple-ID (E-Mail): ").strip()
    pw   = os.environ.get("ICLOUD_PASS") or getpass("App-spezifisches Passwort: ")

    if not os.environ.get("ICLOUD_USER"):
        save = input("Zugangsdaten in .env speichern? (j/n): ").strip().lower()
        if save == "j":
            ENV_FILE.write_text(f"ICLOUD_USER={user}\nICLOUD_PASS={pw}\n")
            print(".env gespeichert.")
    return user, pw


def get_principal_url(session, user):
    """
    Ermittelt die Principal-URL des Nutzers via PROPFIND auf /.
    iCloud leitet dabei auf /[dsnumber]/ weiter.
    """
    body = """<?xml version="1.0" encoding="utf-8"?>
<d:propfind xmlns:d="DAV:">
  <d:prop>
    <d:current-user-principal/>
  </d:prop>
</d:propfind>"""

    r = session.request(
        "PROPFIND",
        f"{CARDDAV_BASE}/",
        data=body.encode(),
        headers={"Depth": "0", "Content-Type": "application/xml"},
    )
    r.raise_for_status()
    root = ET.fromstring(r.text)
    href = root.find(".//d:current-user-principal/d:href", NS)
    if href is None:
        raise RuntimeError("Konnte Principal-URL nicht ermitteln.")
    # iCloud liefert hier teils absolute URLs (eigener Shard-Server),
    # teils relative Pfade — urljoin behandelt beide Fälle korrekt.
    return urljoin(r.url, href.text)


def get_all_addressbook_urls(session, principal_url):
    """
    Ermittelt ALLE Adressbuch-Collections des Nutzers.
    Gibt Liste von (url, displayname) zurück.
    """
    body = """<?xml version="1.0" encoding="utf-8"?>
<d:propfind xmlns:d="DAV:" xmlns:card="urn:ietf:params:xml:ns:carddav">
  <d:prop>
    <card:addressbook-home-set/>
  </d:prop>
</d:propfind>"""

    r = session.request(
        "PROPFIND",
        principal_url,
        data=body.encode(),
        headers={"Depth": "0", "Content-Type": "application/xml"},
    )
    r.raise_for_status()
    root = ET.fromstring(r.text)
    href = root.find(".//card:addressbook-home-set/d:href", NS)
    if href is None:
        raise RuntimeError("Konnte Adressbuch-Home-Set nicht ermitteln.")

    ab_home = urljoin(r.url, href.text)

    # Listet verfügbare Adressbücher auf
    r2 = session.request(
        "PROPFIND",
        ab_home,
        data="""<?xml version="1.0" encoding="utf-8"?>
<d:propfind xmlns:d="DAV:">
  <d:prop><d:displayname/><d:resourcetype/></d:prop>
</d:propfind>""".encode(),
        headers={"Depth": "1", "Content-Type": "application/xml"},
    )
    r2.raise_for_status()
    root2 = ET.fromstring(r2.text)

    books = []
    for resp in root2.findall("d:response", NS):
        rt = resp.find(".//d:resourcetype", NS)
        if rt is not None and rt.find("card:addressbook", NS) is not None:
            h = resp.find("d:href", NS)
            dn = resp.find(".//d:displayname", NS)
            if h is not None:
                name = dn.text if (dn is not None and dn.text) else "(ohne Namen)"
                books.append((urljoin(ab_home, h.text), name))

    if not books:
        raise RuntimeError("Kein Adressbuch gefunden.")
    return books


def get_addressbook_url(session, principal_url):
    """Ermittelt die (erste) Adressbuch-URL – für Import/Löschen."""
    return get_all_addressbook_urls(session, principal_url)[0][0]


def _parse_contact_hrefs(xml_text, addressbook_url):
    """Extrahiert {url: etag} aller Kontakt-Ressourcen aus einer Multistatus-Antwort."""
    found = {}
    root = ET.fromstring(xml_text)
    ab = addressbook_url.rstrip("/")
    for resp in root.findall("d:response", NS):
        rt = resp.find(".//d:resourcetype/d:collection", NS)
        if rt is not None:
            continue  # Überspringe Sammlungen (das Adressbuch selbst)
        href = resp.find("d:href", NS)
        if href is None or not href.text:
            continue
        etag = resp.find(".//d:getetag", NS)
        # Absolute URL aufbauen (iCloud-hrefs können relativ oder absolut sein)
        url = urljoin(addressbook_url, href.text)
        if url.rstrip("/") == ab:
            continue  # das Adressbuch selbst
        found[url] = etag.text if etag is not None else ""
    return found


PROPFIND_HREFS_BODY = """<?xml version="1.0" encoding="utf-8"?>
<d:propfind xmlns:d="DAV:">
  <d:prop>
    <d:getetag/>
    <d:resourcetype/>
  </d:prop>
</d:propfind>"""

ADDRESSBOOK_QUERY_BODY = """<?xml version="1.0" encoding="utf-8"?>
<card:addressbook-query xmlns:d="DAV:" xmlns:card="urn:ietf:params:xml:ns:carddav">
  <d:prop>
    <d:getetag/>
  </d:prop>
  <card:filter>
    <card:prop-filter name="UID"/>
  </card:filter>
</card:addressbook-query>"""

SYNC_COLLECTION_BODY = """<?xml version="1.0" encoding="utf-8"?>
<d:sync-collection xmlns:d="DAV:">
  <d:sync-token></d:sync-token>
  <d:sync-level>1</d:sync-level>
  <d:prop>
    <d:getetag/>
  </d:prop>
</d:sync-collection>"""


def _enumerate(session, addressbook_url, method, body):
    """Führt PROPFIND/REPORT aus und liefert {url: etag}; bei Fehler ({}, Fehlertext)."""
    try:
        r = session.request(
            method,
            addressbook_url,
            data=body.encode(),
            headers={"Depth": "1", "Content-Type": "application/xml"},
        )
        if r.status_code in (200, 207):
            return _parse_contact_hrefs(r.text, addressbook_url), None
        return {}, f"HTTP {r.status_code}"
    except Exception as e:
        return {}, str(e)


def fetch_all_contact_hrefs(session, addressbook_url, verbose=True):
    """
    Gibt Liste von (href, etag) aller Kontakte im Adressbuch zurück.

    iClouds PROPFIND listet in manchen Fällen einzelne Kontakte nicht mit auf.
    Deshalb werden drei Enumerationsverfahren kombiniert und ihre Ergebnisse
    vereinigt (Union kann nur mehr Kontakte finden, nie weniger):
      1. PROPFIND (Depth:1)
      2. addressbook-query REPORT (alle Karten mit UID)
      3. sync-collection REPORT (vollständige Sync-Auflistung, RFC 6578)
    Mit verbose=True wird pro Verfahren die Trefferzahl ausgegeben (Diagnose).
    """
    pf, pf_err = _enumerate(session, addressbook_url, "PROPFIND", PROPFIND_HREFS_BODY)
    aq, aq_err = _enumerate(session, addressbook_url, "REPORT", ADDRESSBOOK_QUERY_BODY)
    sc, sc_err = _enumerate(session, addressbook_url, "REPORT", SYNC_COLLECTION_BODY)

    merged = {}
    for d in (pf, aq, sc):
        for url, etag in d.items():
            merged.setdefault(url, etag)

    if verbose:
        def fmt(name, d, err):
            return f"{name}={len(d)}" + (f" (Fehler: {err})" if err else "")
        print("    Enumeration: " + " | ".join([
            fmt("PROPFIND", pf, pf_err),
            fmt("addressbook-query", aq, aq_err),
            fmt("sync-collection", sc, sc_err),
        ]) + f" | Union={len(merged)}")

    if not merged and pf_err:
        raise RuntimeError(f"Konnte Kontakte nicht auflisten (PROPFIND: {pf_err}).")
    return list(merged.items())


def fetch_vcard(session, url, retries=3):
    """Lädt eine einzelne vCard per GET. Erwartet eine absolute URL.
    Wiederholt bei transienten Fehlern, damit der UID-Abgleich nicht durch
    einzelne Aussetzer Kontakte übersieht (sonst würden sie als 'neu' gelten)."""
    last = None
    for attempt in range(retries):
        try:
            r = session.get(url, headers={"Accept": "text/vcard"}, timeout=30)
            r.raise_for_status()
            return r.text
        except Exception as e:
            last = e
            if attempt < retries - 1:
                time.sleep(0.5 * (attempt + 1))
    raise last


def extract_uid(vcard_text):
    """Extrahiert den UID-Wert aus einer vCard.
    Erst entfalten (RFC-6350-Fortsetzungszeilen zusammenführen) - sonst würden
    lange UIDs (z.B. Outlook-GlobalObjectIds über zwei Zeilen) an der ersten
    physischen Zeile abgeschnitten. iCloud/Thunderbird-Altlasten kleben
    außerdem manchmal eine weitere Property direkt an die UID (z.B.
    UID:...dbbX-ALT-NOTE:...). Solchen Property-Müll schneiden wir ab, sonst
    passt die UID beim Import nicht und der Kontakt würde fälschlich neu
    (dupliziert) angelegt."""
    unfolded = "\n".join(unfold_vcard(vcard_text))
    m = re.search(r"^UID[^:]*:(.+)$", unfolded, re.MULTILINE)
    if not m:
        return None
    uid = m.group(1).strip()
    m2 = re.search(r"[A-Z][A-Z0-9-]*:", uid)  # angehängter Property-Name?
    if m2 and m2.start() > 0:
        uid = uid[:m2.start()]
    return uid.strip()


# ---------------------------------------------------------------------------
# Foto-Einbettung (URI -> Base64)
# ---------------------------------------------------------------------------

def unfold_vcard(text):
    """Hebt vCard-Zeilenfaltung auf (Fortsetzungszeilen beginnen mit Space/Tab)."""
    out = []
    for line in text.splitlines():
        if line[:1] in (" ", "\t") and out:
            out[-1] += line[1:]
        else:
            out.append(line)
    return out


def fold_line(line, limit=73):
    """Faltet eine lange vCard-Zeile (Folding nach RFC 6350: max ~75 Oktette)."""
    if len(line) <= limit:
        return line
    parts = [line[:limit]]
    rest = line[limit:]
    while rest:
        parts.append(" " + rest[:limit - 1])
        rest = rest[limit - 1:]
    return "\r\n".join(parts)


def _photo_subtype(content_type, data):
    """Ermittelt den Bild-Subtyp (JPEG/PNG/...) aus Content-Type oder Magic-Bytes."""
    if content_type:
        ct = content_type.lower()
        if "jpeg" in ct or "jpg" in ct:
            return "JPEG"
        if "png" in ct:
            return "PNG"
        if "gif" in ct:
            return "GIF"
        if "webp" in ct:
            return "WEBP"
    if data[:3] == b"\xff\xd8\xff":
        return "JPEG"
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return "PNG"
    if data[:6] in (b"GIF87a", b"GIF89a"):
        return "GIF"
    return "JPEG"  # iCloud liefert in aller Regel JPEG


_PHOTO_SIZE_WARN_BYTES = 150 * 1024  # iCloud hat groessere Fotos schon mit HTTP 403 abgelehnt


def normalize_photo_type(vcard_text):
    """Korrigiert ein falsch deklariertes PHOTO-TYPE (z.B. TYPE=JPEG bei
    tatsaechlichen PNG-Bytes) anhand der echten Magic-Bytes - iCloud hat so
    eine Karte schon mit HTTP 403 abgelehnt. Warnt zusaetzlich bei sehr
    grossen Fotos, die (auch mit korrektem Typ) an einem iCloud-Limit
    scheitern koennen; eine automatische Verkleinerung braeuchte eine
    Bildbibliothek (Pillow) und ist bewusst nicht eingebaut, damit das
    Skript ohne kompilierte Abhaengigkeiten lauffaehig bleibt (z.B. a-Shell).
    Gibt (vcard_text, warnungen) zurueck."""
    lines = unfold_vcard(vcard_text)
    warnings = []
    for i, l in enumerate(lines):
        if not l.upper().startswith("PHOTO"):
            continue
        head, sep, val = l.partition(":")
        if not sep or val.lower().startswith(("http://", "https://")):
            continue  # URL-Verweis, keine eingebetteten Bilddaten
        try:
            data = base64.b64decode(val)
        except Exception:
            continue
        real_type = _photo_subtype("", data)
        m = re.search(r"TYPE=([A-Za-z0-9]+)", head, re.IGNORECASE)
        declared = m.group(1).upper() if m else None
        if declared and declared != real_type:
            head = re.sub(r"TYPE=[A-Za-z0-9]+", f"TYPE={real_type}", head, flags=re.IGNORECASE)
            lines[i] = f"{head}:{val}"
        if len(data) > _PHOTO_SIZE_WARN_BYTES:
            warnings.append(f"{len(data) / 1024:.0f} KB")
    return "\r\n".join(fold_line(x) for x in lines), warnings


def embed_photos(session, vcard_text):
    """
    Ersetzt PHOTO-Verweise (PHOTO;...;VALUE=uri:https://...) durch eingebettete
    Base64-Daten (PHOTO;ENCODING=b;TYPE=...). Gibt (vcard_text, anzahl) zurück.
    Bei Fehlern bleibt der ursprüngliche Verweis erhalten.
    """
    lines = unfold_vcard(vcard_text)
    embedded = 0
    for idx, line in enumerate(lines):
        if not line.upper().startswith("PHOTO"):
            continue
        head, sep, value = line.partition(":")
        if not sep or not value.lower().startswith(("http://", "https://")):
            continue  # bereits eingebettet oder kein URL-Verweis

        # iCloud faltet bei "geteilten Fotos" die Property X-SHARED-PHOTO-
        # DISPLAY-PREF direkt an die PHOTO-URL an. Saubere Karten-URL bis zum
        # Hash extrahieren, angehängten Property-Müll abschneiden.
        m = re.match(r"https?://\S*?/ck/card/[0-9a-fA-F]+", value.strip())
        photo_url = m.group(0) if m else value.strip().split("X-SHARED-PHOTO-DISPLAY-PREF")[0]

        data = resp = None
        for attempt in range(3):  # kleiner Retry gegen transiente Aussetzer
            try:
                resp = session.get(photo_url, timeout=30)
                resp.raise_for_status()
                data = resp.content
                break
            except Exception as e:
                if attempt == 2:
                    print(f"\n  WARNUNG: Foto konnte nicht geladen werden ({e})")
                else:
                    time.sleep(0.5 * (attempt + 1))
        if not data:
            continue

        subtype = _photo_subtype(resp.headers.get("Content-Type", ""), data)
        b64 = base64.b64encode(data).decode("ascii")

        # Parameter aus dem Kopf übernehmen, VALUE entfernen, ENCODING/TYPE setzen
        params = [p for p in head.split(";")[1:] if not p.upper().startswith("VALUE=")]
        params = [p for p in params if not p.upper().startswith("ENCODING=")
                  and not p.upper().startswith("TYPE=")]
        new_head = ";".join(["PHOTO", *params, "ENCODING=b", f"TYPE={subtype}"])
        lines[idx] = f"{new_head}:{b64}"
        embedded += 1

    # Alle Zeilen RFC-6350-konform falten (Unfold/Rejoin hätte sonst lange
    # iCloud-Felder wie X-ADDRESSING-GRAMMAR ungefaltet gelassen).
    return "\r\n".join(fold_line(l) for l in lines), embedded


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def export_contacts(session, addressbook_urls, embed=True, progress=None):
    """Kernlogik des Exports (ohne Datei-/Konsolenausgabe), damit sowohl die
    CLI als auch die Web-Oberfläche denselben Code nutzen.

    progress(i, total) wird pro geladenem Kontakt aufgerufen (optional).
    Gibt ein Dict zurück: text, total, loaded, photos, warnings.
    """
    if isinstance(addressbook_urls, str):
        addressbook_urls = [(addressbook_urls, "")]

    href_map = {}
    for ab_url, ab_name in addressbook_urls:
        for href, etag in fetch_all_contact_hrefs(session, ab_url):
            href_map.setdefault(href, etag)
    hrefs = list(href_map.items())
    total = len(hrefs)

    photo_count = 0
    vcards = []
    warnings = []
    for i, (href, etag) in enumerate(hrefs, 1):
        if progress:
            progress(i, total)
        try:
            vcard = fetch_vcard(session, href)
            if embed:
                # iCloud liefert Fotos als URL-Verweis statt eingebettet -
                # hier werden sie nachgeladen und als Base64 eingebettet.
                vcard, n = embed_photos(session, vcard)
                photo_count += n
            vcards.append(vcard.strip())
        except Exception as e:
            warnings.append(f"{href} konnte nicht geladen werden: {e}")
        time.sleep(0.05)  # sanftes Rate-Limiting

    return {
        "text": "\r\n".join(vcards) + "\r\n",
        "total": total,
        "loaded": len(vcards),
        "photos": photo_count,
        "warnings": warnings,
    }


def cmd_export(args, session, addressbook_urls):
    output = Path(args.output)
    embed = not getattr(args, "skip_photos", False)

    if isinstance(addressbook_urls, str):
        books = [(addressbook_urls, "")]
    else:
        books = addressbook_urls
    print(f"Durchsuche {len(books)} Adressbuch/Adressbücher ...")

    def progress(i, total):
        print(f"\r  Lade {i}/{total} ...", end="", flush=True)

    result = export_contacts(session, books, embed=embed, progress=progress)

    for w in result["warnings"]:
        print(f"\n  WARNUNG: {w}")
    print(f"\n{result['loaded']} Kontakte geladen.")
    if embed:
        print(f"{result['photos']} Fotos als Base64 eingebettet.")

    output.write_text(result["text"], encoding="utf-8")
    print(f"Gespeichert: {output} ({output.stat().st_size / 1024:.1f} KB)")


# ---------------------------------------------------------------------------
# Excel-Export (VCF -> XLSX) / Excel-Import (XLSX -> VCF)
# ---------------------------------------------------------------------------
#
# Standard-Workflow:
#   export -> to-excel -> (in Excel bearbeiten) -> from-excel -> import --dry-run -> import
#
# Telefone/E-Mails/Adressen/URLs werden in einer Zelle als "Label: Wert"
# zusammengefasst, mehrere Einträge durch " | " getrennt, z.B.:
#   mobil: +49171234567 | Arbeit: +494412345
# Als Label gelten "mobil", "privat", "Arbeit" (feste Bedeutung) - alles
# andere wird als eigenes Apple-Label (X-ABLabel) übernommen.

EXCEL_COLUMNS = [
    "UID", "Löschen", "Anzeigename", "Nachname", "Vorname", "Namenszusatz", "Praefix", "Suffix",
    "Spitzname", "Organisation", "Abteilung", "Titel", "Telefone", "E-Mails",
    "Adressen", "Geburtstag", "URLs", "Notiz", "Kategorien",
]

# Marker-Property in der vCard: markiert einen Kontakt zum Löschen aus iCloud.
# Wird beim Import ausgewertet (DELETE statt PUT) und danach ignoriert.
DELETE_MARKER = "X-VCFSYNC-DELETE"

# Zell-Werte in der Spalte "Löschen", die als "ja, löschen" gelten.
_DELETE_TRUE = {"x", " x", "ja", "j", "1", "true", "wahr", "yes", "y", "löschen"}


def _cell_is_delete(value):
    return str(value or "").strip().lower() in _DELETE_TRUE


def wants_delete(vcard_text):
    """True, wenn die vCard den Lösch-Marker trägt."""
    for l in unfold_vcard(vcard_text):
        head, sep, val = l.partition(":")
        if sep and head.split(";")[0].split(".")[-1].upper() == DELETE_MARKER:
            return val.strip() not in ("", "0", "false")
    return False

_ABLABEL_RE = re.compile(r"_\$!<(.+?)>!\$_")
_TEL_LABELS = {"CELL": "mobil", "HOME": "privat", "WORK": "Arbeit"}
_TEL_IGNORE = {"VOICE", "PREF"}
_EMAIL_LABELS = {"HOME": "privat", "WORK": "Arbeit"}
_EMAIL_IGNORE = {"INTERNET", "PREF"}
_ADR_LABELS = {"HOME": "privat", "WORK": "Arbeit"}


def _decode_ablabel(value):
    """Apples Standard-Label-Konstanten (_$!<Other>!$_ usw.) in Klartext."""
    m = _ABLABEL_RE.match(value or "")
    return m.group(1) if m else (value or "")


def _field_label(segs, item, item_labels, type_map, ignore):
    """Lesbares Label (mobil/privat/Arbeit/eigenes) einer TEL/EMAIL/ADR/URL-
    Property - entweder aus einem itemN.X-ABLabel oder aus TYPE=."""
    if item and item in item_labels:
        return _decode_ablabel(item_labels[item]) or "Sonstige"
    for p in segs[1:]:
        if p.upper().startswith("TYPE="):
            for x in p[5:].split(","):
                u = x.strip().upper()
                if u and u not in ignore:
                    return type_map.get(u, x.strip())
    return "Sonstige"


def vcard_to_fields(card_lines):
    """Wandelt eine (bereits entfaltete) vCard in ein Feld-Dict fürs Excel um."""
    item_labels = {}
    for l in card_lines:
        head = l.partition(":")[0].split(";")[0]
        if "." in head and head.split(".", 1)[1].upper() == "X-ABLABEL":
            item_labels[head.split(".", 1)[0]] = l.partition(":")[2]

    d = {"UID": "", "FN": "", "Nachname": "", "Vorname": "", "Zusatz": "", "Praefix": "",
         "Suffix": "", "Nick": "", "ORG": "", "Abt": "", "Titel": "", "BDAY": "",
         "NOTE": "", "CAT": "", "TEL": [], "EMAIL": [], "ADR": [], "URL": [], "PHOTO_B64": ""}

    for l in card_lines:
        head = l.partition(":")[0]
        val = l.partition(":")[2]
        segs = head.split(";")
        item = segs[0].split(".", 1)[0] if "." in segs[0] else None
        base = segs[0].split(".", 1)[-1].upper()

        if base == "UID":
            d["UID"] = val
        elif base == "FN":
            d["FN"] = val
        elif base == "N":
            p = (val.split(";") + [""] * 5)[:5]
            d["Nachname"], d["Vorname"], d["Zusatz"], d["Praefix"], d["Suffix"] = p
        elif base == "NICKNAME":
            d["Nick"] = val
        elif base == "ORG":
            parts = val.split(";")
            d["ORG"] = parts[0]
            d["Abt"] = parts[1] if len(parts) > 1 else ""
        elif base == "TITLE":
            d["Titel"] = val
        elif base == "BDAY":
            d["BDAY"] = val.replace("value=date:", "")
        elif base == "NOTE":
            d["NOTE"] = val.replace("\\n", " ").replace("\\,", ",")
        elif base == "CATEGORIES":
            d["CAT"] = val.replace("\\,", ",")
        elif base == "TEL":
            label = _field_label(segs, item, item_labels, _TEL_LABELS, _TEL_IGNORE)
            d["TEL"].append(f"{label}: {val.strip()}")
        elif base == "EMAIL":
            label = _field_label(segs, item, item_labels, _EMAIL_LABELS, _EMAIL_IGNORE)
            d["EMAIL"].append(f"{label}: {val.strip()}")
        elif base == "URL":
            label = _field_label(segs, item, item_labels, {}, {"PREF"})
            d["URL"].append(f"{label}: {val.strip()}")
        elif base == "ADR":
            label = _field_label(segs, item, item_labels, _ADR_LABELS, {"PREF"})
            p = (val.split(";") + [""] * 7)[:7]
            street = (p[5] + " " + p[3]).strip()
            flat = " ".join(x for x in [p[2], street, p[4], p[6]] if x.strip())
            flat = flat.replace("\\n", " ").replace("\\,", ",")
            d["ADR"].append(f"{label}: {flat}")
        elif base == "PHOTO" and not val.lower().startswith(("http://", "https://")):
            d["PHOTO_B64"] = val

    return d


# Feldnamen für die Änderungsmeldung (in Anzeige-Reihenfolge)
_COMPARE_LABELS = {
    "Name": "Name", "Spitzname": "Spitzname", "Organisation": "Organisation",
    "Titel": "Titel", "Geburtstag": "Geburtstag", "Notiz": "Notiz",
    "Kategorien": "Kategorien", "Telefone": "Telefone", "E-Mails": "E-Mails",
    "URLs": "URLs", "Adressen": "Adressen",
}


def _semantic_fields(vcard_text):
    """Normalisiert eine vCard auf die inhaltlich relevanten Felder, so dass
    sich zwei Karten vergleichen lassen, ohne dass Nebensächlichkeiten
    (Property-Reihenfolge, iCloud-Serverfelder, unterschiedliche Faltung, doppelt
    auftauchende PLZ nach dem Excel-Rundlauf) als Änderung durchschlagen.
    Fotos werden bewusst ausgeklammert (ändern sich praktisch nie über Excel).
    """
    d = vcard_to_fields(unfold_vcard(vcard_text))

    def as_set(items):
        return tuple(sorted(x.strip() for x in items if x.strip()))

    def adr_tokens(items):
        # "label: Freitext" -> label + alphabetisch sortierte, eindeutige Wörter.
        # Neutralisiert Reihenfolge und die nach dem Rundlauf doppelte PLZ.
        out = []
        for it in items:
            label, _, val = it.partition(":")
            toks = " ".join(sorted(set(val.split())))
            out.append(f"{label.strip().lower()}|{toks}")
        return tuple(sorted(out))

    return {
        "Name": (d["FN"].strip(), d["Nachname"].strip(), d["Vorname"].strip(),
                 d["Zusatz"].strip(), d["Praefix"].strip(), d["Suffix"].strip()),
        "Spitzname": d["Nick"].strip(),
        "Organisation": (d["ORG"].strip(), d["Abt"].strip()),
        "Titel": d["Titel"].strip(),
        "Geburtstag": d["BDAY"].strip(),
        "Notiz": d["NOTE"].strip(),
        "Kategorien": d["CAT"].strip(),
        "Telefone": as_set(d["TEL"]),
        "E-Mails": as_set(d["EMAIL"]),
        "URLs": as_set(d["URL"]),
        "Adressen": adr_tokens(d["ADR"]),
    }


def changed_fields(old_vcard, new_vcard):
    """Gibt die Liste der Felder zurück, die sich zwischen alter (iCloud) und
    neuer (Import) Karte inhaltlich unterscheiden. Leere Liste = keine echte
    Änderung (der PUT würde denselben Inhalt schreiben)."""
    a = _semantic_fields(old_vcard)
    b = _semantic_fields(new_vcard)
    return [_COMPARE_LABELS[k] for k in _COMPARE_LABELS if a.get(k) != b.get(k)]


def cmd_to_excel(args):
    text = Path(args.input).read_text(encoding="utf-8")
    cards = split_vcards(text)

    rows = []
    for card in cards:
        lines = [l for l in unfold_vcard(card) if l.strip()]
        is_group = any(
            l.partition(":")[0].split(";")[0].split(".")[-1].upper() == "X-ADDRESSBOOKSERVER-KIND"
            for l in lines
        )
        if is_group:
            continue  # Gruppen ueberspringen - im iPhone ohnehin nicht sichtbar/bearbeitbar
        rows.append(vcard_to_fields(lines))

    rows.sort(key=lambda d: (d["Nachname"] or d["FN"]).lower())

    chunk = 32000  # unter Excels Zellenlimit von 32.767 Zeichen
    max_chunks = max((math.ceil(len(d["PHOTO_B64"]) / chunk) for d in rows if d["PHOTO_B64"]), default=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kontakte"
    header = EXCEL_COLUMNS + [f"Foto_Base64_{i + 1}" for i in range(max_chunks)]
    ws.append(header)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(header) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for d in rows:
        photo_chunks = [d["PHOTO_B64"][i:i + chunk] for i in range(0, len(d["PHOTO_B64"]), chunk)]
        photo_chunks += [""] * (max_chunks - len(photo_chunks))
        ws.append([
            d["UID"], "", d["FN"], d["Nachname"], d["Vorname"], d["Zusatz"], d["Praefix"], d["Suffix"],
            d["Nick"], d["ORG"], d["Abt"], d["Titel"], " | ".join(d["TEL"]), " | ".join(d["EMAIL"]),
            " | ".join(d["ADR"]), d["BDAY"], " | ".join(d["URL"]), d["NOTE"], d["CAT"],
        ] + photo_chunks)

    ws.freeze_panes = "C2"  # Zeile 1 + UID/Löschen bleiben beim Scrollen sichtbar
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{len(rows) + 1}"
    #        UID Lösch Anz  Nach Vor  Zus Präf Suf Spitz Org  Abt Tit  Tel E-M Adr BDay URL Notiz Kat
    widths = [38, 9,   22,  16,  14,  12, 8,   8,  14,   22,  16, 14,  40, 34, 40, 12,  26, 30,   20]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for i in range(max_chunks):
        ws.column_dimensions[get_column_letter(len(EXCEL_COLUMNS) + 1 + i)].width = 40
    ws.column_dimensions["A"].hidden = True  # UID nicht versehentlich verändern

    wb.save(args.output)
    photo_count = sum(1 for d in rows if d["PHOTO_B64"])
    print(f"{len(rows)} Kontakte geschrieben ({photo_count} mit Foto).")
    print(f"Gespeichert: {args.output}")


def _vesc(value):
    """Escaped Sonderzeichen für vCard-Werte (RFC 6350)."""
    return (value or "").replace("\\", "\\\\").replace(",", "\\,").replace(";", "\\;")


def _parse_label_pairs(cell_value):
    """Zerlegt eine Excel-Zelle im Format 'Label: Wert | Label: Wert' in Paare."""
    out = []
    if not (cell_value or "").strip():
        return out
    for part in cell_value.split(" | "):
        label, _, value = part.partition(": ")
        label, value = label.strip(), value.strip()
        if value:
            out.append((label, value))
    return out


def _tel_property(label, value, next_item):
    lab = label.lower()
    if lab == "mobil":
        return [f"TEL;type=CELL;type=VOICE:{value}"], next_item
    if lab == "privat":
        return [f"TEL;type=HOME;type=VOICE:{value}"], next_item
    if lab == "arbeit":
        return [f"TEL;type=WORK;type=VOICE:{value}"], next_item
    item = f"item{next_item}"
    ablabel = "_$!<Other>!$_" if lab in ("sonstige", "other") else label
    return [f"{item}.TEL;type=VOICE:{value}", f"{item}.X-ABLabel:{ablabel}"], next_item + 1


def _email_property(label, value, next_item):
    lab = label.lower()
    if lab == "privat":
        return [f"EMAIL;type=INTERNET;type=HOME:{value}"], next_item
    if lab == "arbeit":
        return [f"EMAIL;type=INTERNET;type=WORK:{value}"], next_item
    item = f"item{next_item}"
    ablabel = "_$!<Other>!$_" if lab in ("sonstige", "other") else label
    return [f"{item}.EMAIL;type=INTERNET:{value}", f"{item}.X-ABLabel:{ablabel}"], next_item + 1


def _adr_property(label, flat_value, next_item):
    """Baut aus dem Freitext eine ADR-Property. Straße/Ort lassen sich aus dem
    Fließtext nicht zuverlässig trennen - der komplette Text landet in der
    Straßen-Komponente, nur die PLZ wird per Regex herausgezogen."""
    zip_match = re.search(r"\b(\d{5})\b", flat_value)
    plz = zip_match.group(1) if zip_match else ""
    comp = ["", "", _vesc(flat_value), "", "", plz, ""]
    lab = label.lower()
    if lab in ("privat", "arbeit"):
        typ = "WORK" if lab == "arbeit" else "HOME"
        return [f"ADR;type={typ}:" + ";".join(comp)], next_item
    item = f"item{next_item}"
    ablabel = "_$!<Other>!$_" if lab in ("sonstige", "other") else label
    return [f"{item}.ADR:" + ";".join(comp), f"{item}.X-ABLabel:{ablabel}"], next_item + 1


def _url_property(label, value, next_item):
    if label.lower() in ("sonstige", "other", ""):
        return [f"URL;type=OTHER:{value}"], next_item
    item = f"item{next_item}"
    return [f"{item}.URL:{value}", f"{item}.X-ABLabel:{label}"], next_item + 1


def row_to_vcard(row, col_index, photo_cols):
    def g(name):
        i = col_index.get(name)
        v = row[i] if i is not None and i < len(row) else None
        return "" if v is None else str(v).strip()

    uid = g("UID") or str(uuid.uuid4())

    # Zum Löschen markiert? Dann nur eine schlanke Karte mit Marker erzeugen -
    # der Import löscht sie in iCloud (per UID), statt sie zu schreiben.
    if _cell_is_delete(g("Löschen")):
        fn = g("Anzeigename") or " ".join(x for x in (g("Vorname"), g("Nachname")) if x)
        lines = ["BEGIN:VCARD", "VERSION:3.0", f"UID:{uid}",
                 f"FN:{_vesc(fn)}", f"{DELETE_MARKER}:1", "END:VCARD"]
        return "\r\n".join(fold_line(l) for l in lines)

    lines = ["BEGIN:VCARD", "VERSION:3.0", f"UID:{uid}"]

    fn = g("Anzeigename")
    nachname, vorname = g("Nachname"), g("Vorname")
    zusatz, praefix, suffix = g("Namenszusatz"), g("Praefix"), g("Suffix")
    if not fn:
        fn = " ".join(x for x in (vorname, nachname) if x)
    lines.append(f"FN:{_vesc(fn)}")
    lines.append(f"N:{_vesc(nachname)};{_vesc(vorname)};{_vesc(zusatz)};{_vesc(praefix)};{_vesc(suffix)}")

    if g("Spitzname"):
        lines.append(f"NICKNAME:{_vesc(g('Spitzname'))}")
    if g("Organisation") or g("Abteilung"):
        org, abt = _vesc(g("Organisation")), g("Abteilung")
        lines.append(f"ORG:{org};{_vesc(abt)}" if abt else f"ORG:{org}")
    if g("Titel"):
        lines.append(f"TITLE:{_vesc(g('Titel'))}")
    if g("Geburtstag"):
        lines.append(f"BDAY:{g('Geburtstag')}")
    if g("Notiz"):
        lines.append(f"NOTE:{_vesc(g('Notiz'))}")
    if g("Kategorien"):
        lines.append(f"CATEGORIES:{_vesc(g('Kategorien'))}")

    next_item = 1
    for label, value in _parse_label_pairs(g("Telefone")):
        props, next_item = _tel_property(label, value, next_item)
        lines += props
    for label, value in _parse_label_pairs(g("E-Mails")):
        props, next_item = _email_property(label, value, next_item)
        lines += props
    for label, value in _parse_label_pairs(g("Adressen")):
        props, next_item = _adr_property(label, value, next_item)
        lines += props
    for label, value in _parse_label_pairs(g("URLs")):
        props, next_item = _url_property(label, value, next_item)
        lines += props

    photo_b64 = "".join(g(c) for c in photo_cols)
    if photo_b64:
        subtype = "JPEG"
        try:
            subtype = _photo_subtype("", base64.b64decode(photo_b64))
        except Exception:
            pass
        lines.append(f"PHOTO;ENCODING=b;TYPE={subtype}:{photo_b64}")

    lines.append("END:VCARD")
    return "\r\n".join(fold_line(l) for l in lines)


def cmd_from_excel(args):
    wb = load_workbook(args.input, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h) if h is not None else "" for h in rows[0]]
    col_index = {h: i for i, h in enumerate(header)}
    photo_cols = [h for h in header if h.startswith("Foto_Base64_")]

    # "Löschen" ist optional - alte Excel-Dateien (ohne die Spalte) bleiben gültig.
    required = [c for c in EXCEL_COLUMNS if c != "Löschen"]
    missing = [c for c in required if c not in col_index]
    if missing:
        print(f"Fehlende Spalten in {args.input}: {', '.join(missing)}")
        print("Wurde die Datei mit 'to-excel' erzeugt und nicht umbenannt/umsortiert?")
        sys.exit(1)

    del_col = col_index.get("Löschen")

    cards = []
    skipped = 0
    marked_delete = 0
    for row in rows[1:]:
        if row is None or all(v is None for v in row):
            continue
        is_delete = del_col is not None and _cell_is_delete(row[del_col] if del_col < len(row) else None)
        has_name = any(
            str(row[col_index[c]] or "").strip()
            for c in ("Anzeigename", "Nachname", "Vorname")
        )
        if not has_name and not is_delete:
            skipped += 1
            continue
        cards.append(row_to_vcard(row, col_index, photo_cols))
        if is_delete:
            marked_delete += 1

    Path(args.output).write_text("\r\n".join(cards) + "\r\n", encoding="utf-8")
    parts = [f"{len(cards)} Kontakte geschrieben"]
    if marked_delete:
        parts.append(f"davon {marked_delete} zum Löschen markiert")
    if skipped:
        parts.append(f"{skipped} leere Zeile(n) übersprungen")
    print(", ".join(parts) + ".")
    print(f"Gespeichert: {args.output}")


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

def split_vcards(text):
    """Zerlegt eine VCF-Datei in einzelne vCard-Strings."""
    cards = []
    current = []
    for line in text.splitlines():
        current.append(line)
        if line.strip().upper() == "END:VCARD":
            cards.append("\r\n".join(current))
            current = []
    return cards


def fetch_existing_uids(session, addressbook_urls, content_out=None):
    """
    Gibt Dict {uid: href} aller vorhandenen Kontakte zurück.
    Lädt dazu jede vCard und liest die UID.

    addressbook_urls kann eine einzelne URL (str) oder eine Liste von
    (url, name)-Tupeln sein. Wie beim Export werden ALLE Adressbücher
    durchsucht - sonst gelten Kontakte aus einem zweiten Adressbuch beim
    Import fälschlich als "neu" (Duplikat-Risiko).

    Wird ein Dict content_out übergeben, wird es mit {uid: vcard_text} gefüllt -
    die Karten werden ohnehin geladen, das kostet keine zusätzlichen Zugriffe.
    Damit lässt sich beim Import feststellen, welche Kontakte sich wirklich
    ändern (statt nur "vorhanden, wird überschrieben").
    """
    if isinstance(addressbook_urls, str):
        addressbook_urls = [(addressbook_urls, "")]

    print("Lese vorhandene Kontakte aus iCloud (für UID-Abgleich) ...")
    href_map = {}
    for ab_url, ab_name in addressbook_urls:
        for href, etag in fetch_all_contact_hrefs(session, ab_url):
            href_map.setdefault(href, etag)
    hrefs = list(href_map.items())

    uid_map = {}
    for i, (href, _) in enumerate(hrefs, 1):
        print(f"\r  {i}/{len(hrefs)} ...", end="", flush=True)
        try:
            vcard = fetch_vcard(session, href)
            uid = extract_uid(vcard)
            if uid:
                uid_map[uid] = href
                if content_out is not None:
                    content_out[uid] = vcard
        except Exception as e:
            print(f"\n  WARNUNG: {href}: {e}")
        time.sleep(0.05)
    print(f"\n{len(uid_map)} UIDs eingelesen.")
    return uid_map


def put_vcard(session, url, vcard_text, etag=None):
    """Schreibt eine vCard per PUT (Update oder Neuanlage)."""
    headers = {
        "Content-Type": "text/vcard; charset=utf-8",
    }
    if etag:
        headers["If-Match"] = etag  # Optimistic Locking bei Updates
    r = session.put(url, data=vcard_text.encode("utf-8"), headers=headers)
    return r


def import_vcards(session, books, primary_url, vcards, dry_run=False, progress=None):
    """Kernlogik des Imports (ohne Datei-/Konsolenausgabe), gemeinsam genutzt
    von CLI und Web-Oberfläche.

    progress(i, total, message) wird pro Kontakt aufgerufen (optional); message
    ist die fertig formatierte Statuszeile (z.B. "[3] AKTUALISIERT: ...").
    Gibt ein Dict zurück: updated, created, errors, total, log.
    """
    # Vorhandene UIDs aus ALLEN Adressbüchern laden (nicht nur dem ersten -
    # sonst gelten Kontakte aus einem zweiten Adressbuch fälschlich als neu).
    # Karteninhalte gleich mit erfassen, um echte Änderungen zu erkennen.
    existing_content = {}
    uid_map = fetch_existing_uids(session, books, content_out=existing_content)

    updated = created = deleted = errors = changed = 0
    changed_list = []
    log = []
    total = len(vcards)

    def emit(i, message):
        log.append(message)
        if progress:
            progress(i, total, message)

    def fn_of(vcard_text):
        m = re.search(r"^FN:(.+)$", "\n".join(unfold_vcard(vcard_text)), re.MULTILINE)
        return m.group(1).strip() if m else ""

    for i, vcard in enumerate(vcards, 1):
        vcard, photo_warnings = normalize_photo_type(vcard)
        for w in photo_warnings:
            emit(i, f"[{i}] Hinweis: Foto ist {w} groß - iCloud hat solche Fotos schon mit HTTP 403 abgelehnt")

        uid = extract_uid(vcard)
        if not uid:
            emit(i, f"[{i}] KEIN UID - übersprungen")
            errors += 1
            continue

        # Zum Löschen markiert: Kontakt aus iCloud entfernen (DELETE statt PUT)
        if wants_delete(vcard):
            if uid not in uid_map:
                emit(i, f"[{i}] LÖSCHEN übersprungen (nicht in iCloud): {uid[:20]}...")
                continue
            url = uid_map[uid]
            if dry_run:
                emit(i, f"[{i}] (dry-run) WÜRDE LÖSCHEN: {uid[:20]}...")
                deleted += 1
                continue
            r = delete_contact(session, url)
            if r.status_code in (200, 204):
                emit(i, f"[{i}] GELÖSCHT: {uid[:20]}...")
                deleted += 1
            else:
                emit(i, f"[{i}] FEHLER Löschen {r.status_code}: {uid[:20]}")
                errors += 1
            time.sleep(0.1)
            continue

        if uid in uid_map:
            # Update: vorhandener Kontakt (uid_map enthält bereits absolute URLs)
            url = uid_map[uid]
            old = existing_content.get(uid)
            fields = changed_fields(old, vcard) if old is not None else None
            if fields:  # unterscheidet sich wirklich vom iCloud-Stand
                changed += 1
                changed_list.append({"name": fn_of(vcard) or uid[:12], "fields": fields})
            # Kennzeichnung in der Statuszeile: geändert / unverändert / unbekannt
            if fields:
                tag = f" [geändert: {', '.join(fields)}]"
            elif fields == []:
                tag = " [unverändert]"
            else:
                tag = ""
            if dry_run:
                emit(i, f"[{i}] (dry-run) WÜRDE AKTUALISIEREN{tag}: {uid[:20]}...")
                updated += 1
                continue
            r = put_vcard(session, url, vcard)
            if r.status_code in (200, 201, 204):
                emit(i, f"[{i}] AKTUALISIERT{tag}: {uid[:20]}...")
                updated += 1
            else:
                emit(i, f"[{i}] FEHLER Update {r.status_code}: {uid[:20]}")
                errors += 1
        else:
            # Neuanlage: neue .vcf-Ressource im Adressbuch
            safe_uid = re.sub(r"[^a-zA-Z0-9\-]", "", uid)
            url = primary_url.rstrip("/") + f"/{safe_uid}.vcf"
            if dry_run:
                emit(i, f"[{i}] (dry-run) WÜRDE NEU ANLEGEN: {uid[:20]}...")
                created += 1
                continue
            r = put_vcard(session, url, vcard)
            if r.status_code in (200, 201, 204):
                emit(i, f"[{i}] NEU ANGELEGT: {uid[:20]}...")
                created += 1
            else:
                emit(i, f"[{i}] FEHLER Neuanlage {r.status_code}: {uid[:20]}")
                errors += 1

        if not dry_run:
            time.sleep(0.1)  # sanftes Rate-Limiting

    return {"updated": updated, "created": created, "deleted": deleted,
            "changed": changed, "changed_list": changed_list,
            "errors": errors, "total": total, "log": log}


def cmd_import(args, session, books, primary_url):
    input_file = Path(args.input)
    if not input_file.exists():
        print(f"Datei nicht gefunden: {input_file}")
        sys.exit(1)

    vcards = split_vcards(input_file.read_text(encoding="utf-8"))
    print(f"{len(vcards)} Kontakte in {input_file} gefunden.")

    dry_run = getattr(args, "dry_run", False)
    result = import_vcards(session, books, primary_url, vcards, dry_run=dry_run,
                           progress=lambda i, total, message: print(f"  {message}"))

    print(f"\nFertig: {result['updated']} aktualisiert "
          f"(davon {result['changed']} inhaltlich geändert), "
          f"{result['created']} neu, {result['deleted']} gelöscht, "
          f"{result['errors']} Fehler.")
    if result["changed_list"]:
        print("\nGeänderte Kontakte:")
        for c in result["changed_list"]:
            print(f"  - {c['name']}: {', '.join(c['fields'])}")


# ---------------------------------------------------------------------------
# Löschen
# ---------------------------------------------------------------------------

def delete_contact(session, url):
    """Löscht eine Kontakt-Ressource per HTTP DELETE."""
    return session.delete(url)


def cmd_delete(args, session, books):
    """Löscht Kontakte anhand ihrer UID (aus --uid und/oder --input VCF)."""
    uids = []
    if getattr(args, "input", None):
        input_file = Path(args.input)
        if not input_file.exists():
            print(f"Datei nicht gefunden: {input_file}")
            sys.exit(1)
        for vcard in split_vcards(input_file.read_text(encoding="utf-8")):
            u = extract_uid(vcard)
            if u:
                uids.append(u)
    for u in getattr(args, "uid", None) or []:
        uids.append(u.strip())
    uids = list(dict.fromkeys(uids))  # Reihenfolge erhalten, Duplikate raus

    if not uids:
        print("Keine UIDs angegeben. Nutze --input <vcf> und/oder --uid <UID>.")
        sys.exit(1)
    print(f"{len(uids)} zu löschende Kontakt-UID(s).")

    uid_map = fetch_existing_uids(session, books)
    dry_run = getattr(args, "dry_run", False)

    deleted = missing = errors = 0
    for i, uid in enumerate(uids, 1):
        if uid not in uid_map:
            print(f"  [{i}] NICHT GEFUNDEN (evtl. schon gelöscht): {uid[:20]}...")
            missing += 1
            continue
        url = uid_map[uid]
        if dry_run:
            print(f"  [{i}] (dry-run) WÜRDE LÖSCHEN: {uid[:20]}...")
            deleted += 1
            continue
        r = delete_contact(session, url)
        if r.status_code in (200, 204):
            print(f"  [{i}] GELÖSCHT: {uid[:20]}...")
            deleted += 1
        else:
            print(f"  [{i}] FEHLER Löschen {r.status_code}: {uid[:20]}")
            errors += 1
        time.sleep(0.1)  # sanftes Rate-Limiting

    print(f"\nFertig: {deleted} gelöscht, {missing} nicht gefunden, {errors} Fehler.")


# ---------------------------------------------------------------------------
# Hauptprogramm
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="iCloud Kontakte via CardDAV exportieren und importieren"
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    exp = sub.add_parser("export", help="Kontakte aus iCloud exportieren")
    exp.add_argument("--output", default="icloud_kontakte.vcf",
                     help="Zieldatei (Standard: icloud_kontakte.vcf)")
    exp.add_argument("--skip-photos", action="store_true",
                     help="Fotos nicht nachladen/einbetten (nur URL-Verweise, schneller)")

    imp = sub.add_parser("import", help="Bearbeitete VCF in iCloud zurückspielen")
    imp.add_argument("--input", required=True,
                     help="Zu importierende VCF-Datei")
    imp.add_argument("--dry-run", action="store_true",
                     help="Nur simulieren, nichts schreiben")

    dele = sub.add_parser("delete", help="Kontakte aus iCloud löschen (per UID oder VCF)")
    dele.add_argument("--input",
                      help="VCF-Datei; alle darin enthaltenen UIDs werden gelöscht")
    dele.add_argument("--uid", action="append",
                      help="Einzelne UID zum Löschen (mehrfach angebbar)")
    dele.add_argument("--dry-run", action="store_true",
                      help="Nur simulieren, nichts löschen")

    toexcel = sub.add_parser("to-excel", help="VCF in eine Excel-Liste zum Bearbeiten wandeln")
    toexcel.add_argument("--input", required=True, help="Quell-VCF-Datei")
    toexcel.add_argument("--output", default="kontakte.xlsx",
                         help="Ziel-Excel-Datei (Standard: kontakte.xlsx)")

    fromexcel = sub.add_parser("from-excel", help="Bearbeitete Excel-Liste zurück in VCF wandeln")
    fromexcel.add_argument("--input", required=True, help="Quell-Excel-Datei")
    fromexcel.add_argument("--output", default="kontakte_bearbeitet.vcf",
                           help="Ziel-VCF-Datei (Standard: kontakte_bearbeitet.vcf)")

    args = parser.parse_args()

    # to-excel/from-excel arbeiten rein lokal auf Dateien - keine iCloud-Verbindung nötig
    if args.cmd == "to-excel":
        cmd_to_excel(args)
        return
    if args.cmd == "from-excel":
        cmd_from_excel(args)
        return

    # Session aufbauen
    user, pw = get_credentials()
    session = requests.Session()
    session.auth = HTTPBasicAuth(user, pw)
    session.headers.update({"User-Agent": "iCloudContactsScript/1.0"})

    print("Verbinde mit iCloud ...")
    try:
        principal_url  = get_principal_url(session, user)
        books = get_all_addressbook_urls(session, principal_url)
        print(f"{len(books)} Adressbuch/Adressbücher gefunden:")
        for url, name in books:
            print(f"  - {name}: {url}")
    except Exception as e:
        print(f"Verbindungsfehler: {e}")
        print("Hinweis: Stelle sicher, dass du ein app-spezifisches Passwort verwendest")
        print("         (appleid.apple.com > Sicherheit > App-spezifische Passwörter)")
        sys.exit(1)

    primary_url = books[0][0]  # erstes Adressbuch für Import/Neuanlage/Löschen
    if args.cmd == "export":
        cmd_export(args, session, books)
    elif args.cmd == "import":
        if getattr(args, "dry_run", False):
            print("DRY-RUN: Es werden keine Daten geschrieben.")
        cmd_import(args, session, books, primary_url)
    elif args.cmd == "delete":
        if getattr(args, "dry_run", False):
            print("DRY-RUN: Es werden keine Daten gelöscht.")
        cmd_delete(args, session, books)


if __name__ == "__main__":
    main()
